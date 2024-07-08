import requests
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

def get_vacancies(vacancy, area, page):
    url = "https://api.hh.ru/vacancies"
    params = {
        "text": vacancy,
        "area": area,  # Specify the desired area ID (1 is Moscow)
        "per_page": 100,  # Number of vacancies per page
        "page": page
    }
    headers = {
        "User-Agent": "Your User Agent",  # Replace with your User-Agent header
    }

    response = requests.get(url, params=params, headers=headers)

    if response.status_code == 200:
        data = response.json()
        vacancies = data.get("items", [])

        # Create a JSON file and write the data
        with open('data/raw/vacancies.json', 'w', encoding='utf-8') as file:
            json.dump(vacancies, file, ensure_ascii=False, indent=4)

        print("Vacancies saved to vacancies.json")
        create_xl_table("data/raw/vacancies.json")
    else:
        print(f"Request failed with status code: {response.status_code}")


def create_xl_table(json_file):
    # Открываем JSON файл
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    try:
        # Пытаемся загрузить существующий XL файл
        workbook = load_workbook('data/raw/vacancies.xlsx')
        worksheet = workbook.active
    except FileNotFoundError:
        # Если файл не существует, создаем новый
        workbook = Workbook()
        worksheet = workbook.active
        # Определяем заголовки столбцов только если файл новый
        headers = ['id', 'name', 'area', 'salary', 'valuta', 'alternate_url', 'employer_name', 'snippet_requirement', 'snippet_responsibility', 'schedule', 'working_time_intervals', 'professional_roles', 'experience', 'employment', 'published_at']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

    # Находим номер строки для начала записи данных (последняя_записанная_строка + 1)
    row = worksheet.max_row + 1

    for vacancy in data:
        worksheet.cell(row, 1, vacancy['id'])
        worksheet.cell(row, 2, vacancy['name'])
        worksheet.cell(row, 3, vacancy['area']['name'])
        if vacancy['salary']:
            salary_from_or_to = vacancy['salary']['from'] if vacancy['salary']['from'] else vacancy['salary']['to']
            worksheet.cell(row, 4, salary_from_or_to or 'None')
            worksheet.cell(row, 5, vacancy['salary']['currency'] or 'None')
        else:
            worksheet.cell(row, 4, 'None')
            worksheet.cell(row, 5, 'None')
        worksheet.cell(row, 6, vacancy['alternate_url'])
        worksheet.cell(row, 7, vacancy['employer']['name'])
        worksheet.cell(row, 8, vacancy['snippet']['requirement'])
        worksheet.cell(row, 9, vacancy['snippet']['responsibility'])
        worksheet.cell(row, 10, vacancy['schedule']['name'])
        working_time_intervals = ', '.join([interval['name'] for interval in vacancy['working_time_intervals']]) if vacancy['working_time_intervals'] else 'None'
        worksheet.cell(row, 11, working_time_intervals)
        professional_roles = ', '.join([role['name'] for role in vacancy['professional_roles']])
        worksheet.cell(row, 12, professional_roles)
        worksheet.cell(row, 13, vacancy['experience']['name'])
        worksheet.cell(row, 14, vacancy['employment']['name']) 
        worksheet.cell(row, 15, vacancy['published_at']) 
        row += 1

    # Сохраняем изменения в XL файл
    workbook.save('data/raw/vacancies.xlsx')
    print("Vacancies saved to vacancies.xls")


#Example usage
def main():
    """
    Реализация автоматического парсинга по всем ключивым словам
    vacancies = [
        'BI Developer', 'Business Development Manager', 'Community Manager', 'Computer vision',
        'Data Analyst', 'Data Engineer', 'Data Science', 'Data Scientist', 'ML Engineer',
        'Machine Learning Engineer', 'ML OPS инженер', 'ML-разработчик', 'Machine Learning',
        'Product Manager', 'Python Developer', 'Web Analyst', 'Аналитик данных',
        'Бизнес-аналитик', 'Веб-аналитик', 'Системный аналитик', 'Финансовый аналитик'
    ]
    for vacancy in vacancies:
        for i in range(20):
            get_vacancies(vacancy, 1, i)
    """
    vacancies = [
        'Аналитик данных', 'Бизнес-аналитик','Финансовый аналитик'
    ]
    for vacancy in vacancies:
        for i in range(20):
            get_vacancies(vacancy, 1, i)
    
if __name__ == "__main__":
    main()